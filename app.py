import streamlit as st
import pandas as pd
import os
import re
from datetime import datetime, date, timedelta
import time 
import numpy as np 
import plotly.express as px
import json
# --- THƯ VIỆN ĐẶC BIỆT CHO EXCEL: XỬ LÝ MERGED CELLS ---
from openpyxl import load_workbook 
# -----------------------
# SỬ DỤNG LITELLM ĐỂ FIX LỖI API INCOMPATIBILITY
from litellm import completion
# CẬP NHẬT: TĂNG THỜI GIAN CHỜ VÀ RETRY CHO ỔN ĐỊNH
from tenacity import retry, stop_after_attempt, wait_random_exponential, retry_if_exception_type

# =============================================================================
# 0. KHỞI TẠO CẤU HÌNH GLOBAL & AI CLIENT (V7.33.4 - CLOUD READY)
# =============================================================================
# --- ĐỌC API KEY TỪ SECRETS HOẶC ENV (KHÔNG DÙNG FILE API_KEY.TXT NỮA) ---
AI_CLIENT_STATUS = False
AI_ERROR = None
AI_MODEL = "openai/gpt-4o-mini" 

try:
    # 1. Thử đọc key từ Streamlit Secrets (Cloud)
    if 'OPENAI_API_KEY' in st.secrets:
        api_key = st.secrets['OPENAI_API_KEY']
    # 2. Thử đọc key từ Biến môi trường (Local)
    else:
        api_key = os.environ.get("OPENAI_API_KEY")

    if api_key:
        if not api_key.startswith(('sk-', 'sk-proj-')):
             AI_ERROR = "❌ Lỗi: API Key có định dạng sai."
        else:
             # Đặt key vào môi trường để litellm sử dụng
             os.environ["OPENAI_API_KEY"] = api_key
             AI_CLIENT_STATUS = True
    else:
        AI_ERROR = "⚠️ Lỗi: Không tìm thấy OPENAI_API_KEY trong Secrets hoặc Biến môi trường."

except Exception as e:
    AI_ERROR = f"❌ Lỗi cấu hình API Key: {e}"
# -------------------------------------------------------------------------

# --- CẤU HÌNH CỘT LINK VIDEO ---
DEFAULT_MENU_VIDEO = {
    "LINK NIỀM TIN": "https://www.youtube.com/watch?v=PoUWP--0CDU",        
    "LINK IUL": "https://www.youtube.com/watch?v=YqL7qMa1PCU&list=PLFkppJwxKoxXNFfYDwntyTQB9JT8tZ0yR",       
    "LINK BỒI THƯỜNG": "https://www.youtube.com/watch?v=XdwWH2bBvnU",      
    "LINK REVIEW KH": "https://www.youtube.com/watch?v=3KWj3A4S-RA"        
}

def load_menu_config():
    config_file = "GUS_CONFIG.TXT"
    menu = DEFAULT_MENU_VIDEO.copy()
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                lines = [line.strip() for line in f if line.strip() and not line.startswith('#')]
            if len(lines) >= 1:
                menu = {}
                for line in lines:
                    if '|' in line:
                        t, u = line.split('|', 1)
                        menu[t.strip()] = u.strip()
                if len(menu) == 0: 
                     menu = DEFAULT_MENU_VIDEO
        except: pass
    return menu

MENU_VIDEO = load_menu_config()
VIDEO_MENU_KEYS = list(MENU_VIDEO.keys()) 

# --- HỆ THỐNG PHÂN TÍCH TỰ ĐỘNG CŨ (Chỉ là fallback/mặc định) ---
STATUS_OPTIONS = [
    "Done (100%)", "Hot Interest (85%)", "Interest (75%)", 
    "Follow Up (50%)", "Unidentified (10%)", "Cold (5%)", "Stop (0%)"
]
STATUS_RULES = [
    ("Stop (0%)", ["từ chối", "ko mua", "dnc", "stop", "sai số", "agent", "block", "thái độ tệ", "phá đám"]),
    ("Done (100%)", ["chốt", "ký đơn", "sold", "paid", "hoàn tất", "đã chốt", "đã lấy full thông tin"]),
    ("Hot Interest (85%)", ["báo giá", "quote", "ssn", "chạy giá", "gửi form", "rất quan tâm", "hứng thú", "đã tư vấn đầy đủ", "đã run quote", "rất nhiệt huyết", "lịch hẹn lấy thông tin"]),
    ("Interest (75%)", ["quan tâm", "muốn tìm hiểu", "coi video", "xem clip", "thể hiện sự quan tâm rõ ràng", "khách quen giới thiệu", "khả năng tham gia cao"]),
    ("Follow Up (50%)", ["gọi lại", "sẽ gọi", "hẹn", "bận", "chưa rảnh", "có tiềm năng", "follow lâu dài", "1-6 tháng"]),
    ("Cold (5%)", ["nghĩ lại", "chưa vội", "ko tiền", "hết tiền", "bó tay", "mua với bên khác", "trốn tìm", "bệnh", "già"]),
    ("Unidentified (10%)", ["none", "rỗng", "chưa tương tác", "ko note", "chưa xác định được ý định", "nhu cầu của khách rõ ràng"]),
]
MAPPING_DICT = {
    "NAME": ["tên", "họ tên", "full name", "fullname", "khách hàng", "tên khách", "lead name", "lead"],
    "Cellphone": ["sđt", "số điện thoại", "phone", "mobile", "tel", "cell", "phone number", "số đt"],
    "Số Tiệm": ["số tiệm", "số phone tiệm", "shop phone", "store phone"], 
    "NOTE": ["ghi chú", "note", "nội dung", "mô tả", "comment", "notes"],
    "Status": ["trạng thái", "tình trạng", "status", "state", "STATUS"], 
    "ASSIGNED": ["sale", "người phụ trách", "nhân viên", "assign to"],
}


# =============================================================================
# 1. CẤU HÌNH GIAO DIỆN & CSS
# =============================================================================
# --- Đặt theme mặc định là light và tiêu đề ---
st.set_page_config(
    page_title="3M-Gus", # ĐÃ ĐỔI TÊN Ở ĐÂY
    page_icon="💎",
    layout="wide",
    initial_sidebar_state="expanded" 
)

# --- HÀM LƯU DỮ LIỆU ĐÃ CHỈNH SỬA ---
def save_dataframe_changes(df_to_save):
    cols_to_remove = [
        "CALL_LINK", "CLEAN_PHONE", 
        "ID", "EDIT", "Cellphone_Link", "Số Tiệm_Link", "CLEAN_SHOP_PHONE",
        "STATUS_SHORT", "TAM_LY_SHORT",
        "VIDEO_GUIDE" 
    ]
    
    df_clean = df_to_save.copy()
    
    if 'LAST_CONTACT_DATE' in df_clean.columns:
         df_clean['LAST_CONTACT_DATE'] = pd.to_datetime(df_clean['LAST_CONTACT_DATE'], errors='coerce').dt.date
    if 'LAST_CALL_DATETIME' in df_clean.columns:
         df_clean['LAST_CALL_DATETIME'] = pd.to_datetime(df_clean['LAST_CALL_DATETIME'], errors='coerce')

    df_clean = df_clean.drop(columns=[col for col in cols_to_remove if col in df_clean.columns], errors='ignore')
    
    # LOẠI BỎ CÁC CỘT LINK VIDEO CŨ KHI LƯU
    df_clean = df_clean.drop(columns=[col for col in VIDEO_MENU_KEYS if col in df_clean.columns], errors='ignore')
    
    # --- KHU VỰC LƯU FILE CÓ THAY ĐỔI ---
    TEMP_FILE = "temp_data.xlsx"
    TARGET_FILE = "data.xlsx"
    MAX_RETRIES = 5

    try:
        if 'Status' not in df_clean.columns and 'STATUS' in df_clean.columns:
             df_clean.rename(columns={'STATUS': 'Status'}, inplace=True)
             
        # Ghi file ra bộ nhớ tạm
        df_clean.to_excel(TEMP_FILE, index=False, engine="openpyxl")
        st.toast("✅ Đã lưu dữ liệu thô vào file tạm thành công!", icon="💾")
        
        # Đổi tên file để ghi đè lên file data.xlsx gốc (FIX lỗi khóa file)
        for attempt in range(MAX_RETRIES):
            try:
                if os.path.exists(TARGET_FILE):
                    os.remove(TARGET_FILE) 
                os.rename(TEMP_FILE, TARGET_FILE) 
                st.toast("✅ Cập nhật file data.xlsx hoàn tất!", icon="💾")
                return 
            except PermissionError as pe:
                if attempt < MAX_RETRIES - 1:
                    st.warning(f"⚠️ Lỗi khóa file. Thử lại sau {2 ** attempt} giây. Vui lòng đóng Excel! ({attempt+1}/{MAX_RETRIES})")
                    time.sleep(2 ** attempt)
                else:
                    raise pe 
            except Exception as e:
                raise e
            
    except Exception as e:
        st.error(f"❌ Lỗi CRITICAL khi lưu file data.xlsx: {e}. Vui lòng đóng file data.xlsx nếu đang mở.")
    
    finally:
        if os.path.exists(TEMP_FILE):
            os.remove(TEMP_FILE) 

# --- CSS TÙY CHỈNH (V7.33.4) ---
st.markdown("""
<style>
    /* Ẩn các thành phần thừa */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    
    /* 0. FORCE LIGHT MODE & TEXT COLOR */
    :root {
        --base-background-color: #FAFAFA !important; 
        --secondary-background-color: #FFFFFF !important; 
        --text-color: #000000 !important;
    }
    /* Ghi đè Dark Mode chính */
    .stApp, .main .block-container {
        background-color: #FAFAFA !important; 
        color: #000000 !important;
    }

    /* 1. CUSTOM SIDEBAR COLOR (Màu Cam/Nâu) */
    section[data-testid="stSidebar"] {
        min-width: 250px !important; 
        background: linear-gradient(180deg, #D35400 0%, #E67E22 100%) !important; 
        color: white !important;
    }
    
    /* 2. FORCE TEXT BLACK (Cho khu vực chính) */
    .stApp, .stMarkdown, .stText, p, h1, h2, h3, div, span, label, div[data-testid="stExpander"] div {
        color: #000000 !important;
    }
    /* Đảm bảo chữ trong sidebar vẫn là trắng */
    section[data-testid="stSidebar"] *, section[data-testid="stSidebar"] h3 {
        color: #FFFFFF !important; 
    }
    
    /* --- FIX: MÀU CHỮ CỦA NÚT LINK BUTTON VIDEO --- */
    /* Target Link Button loại primary trong sidebar (Link Video) */
    section[data-testid="stSidebar"] div[data-testid="stLinkButton"] button[kind="primary"] {
        background-color: #D35400 !important; 
        color: #FFFFFF !important; 
        border: 1px solid #FFFFFF !important; 
    }
    /* Buộc chữ bên trong nút Link Button primary thành màu TRẮNG */
    section[data-testid="stSidebar"] div[data-testid="stLinkButton"] button[kind="primary"] * {
        color: #FFFFFF !important; 
    }
    /* BẮT LẤY SVG ICON */
    section[data-testid="stSidebar"] div[data-testid="stLinkButton"] button[kind="primary"] svg {
        fill: #FFFFFF !important;
    }


    /* --- 3. INPUT/TEXTAREA FIX (KHU VỰC CHÍNH) --- */
    
    /* Selector Input/Textarea - MÀU CHỮ */
    div[data-testid="stTextInput"] input,
    div[data-testid="stTextInput"] textarea,
    div[data-testid="stSelectbox"] input 
    {
        background-color: #EBF5FB !important; /* Nền xanh băng nhạt */
        border: 1px solid #B0C4DE !important; 
        color: #000000 !important;
        -webkit-text-fill-color: #000000 !important; 
    }
    /* Placeholder (Chữ mờ khi chưa nhập) */
    div[data-testid="stTextInput"] input::placeholder,
    div[data-testid="stTextInput"] textarea::placeholder {
        color: #4C4C4C !important; /* Xám đen nhẹ */
        opacity: 1 !important;
    }
    
    /* Selectbox/Dropdown (Lọc dữ liệu) */
    div[data-testid="stSelectbox"] div[data-baseweb="select"] {
        background-color: #EBF5FB !important; 
        border: 1px solid #B0C4DE !important; 
        color: #000000 !important;
    }
    div[data-testid="stSelectbox"] > div[data-baseweb="select"] > div:first-child { 
         background-color: #EBF5FB !important; 
         color: #000000 !important;
    }
    
    /* --- ULTIMATE INPUT/TEXTAREA FIX (SIDEBAR) --- */
    /* Selectbox trong Sidebar (Phần Chọn Khách Hàng) */
    section[data-testid="stSidebar"] div[data-testid="stSelectbox"] div[data-baseweb="select"] {
        background-color: #FFFFFF !important; 
        color: #000000 !important;
    }
    /* **MÀU CHỮ TRONG VÙNG SELECTBOX CHÍNH CỦA SIDEBAR** */
    section[data-testid="stSidebar"] div[data-testid="stSelectbox"] div[data-baseweb="select"] div[data-testid="stPlaceholder"],
    section[data-testid="stSidebar"] div[data-testid="stSelectbox"] div[data-baseweb="select"] span,
    section[data-testid="stSidebar"] div[data-testid="stSelectbox"] input,
    section[data-testid="stSidebar"] .stSelectbox .st-bh, 
    section[data-testid="stSidebar"] .stSelectbox .st-br,
    section[data-testid="stSidebar"] .stSelectbox .st-bu 
    {
        color: #000000 !important; 
        -webkit-text-fill-color: #000000 !important; 
    }
    
    /* FIX: TEXT AREA/INPUT TRONG SIDEBAR (KHU VỰC KỊCH BẢN) */
    section[data-testid="stSidebar"] div[data-testid="stTextarea"] textarea,
    section[data-testid="stSidebar"] div[data-testid="stTextInput"] input
    {
        background-color: #FFFFFF !important; /* Nền trắng */
        color: #000000 !important; /* Chữ đen */
        -webkit-text-fill-color: #000000 !important; 
    }
    
    /* Các tùy chọn trong dropdown list (đảm bảo không bị tối) */
    .stSelectbox div[data-baseweb="select"] div[role="option"] {
        background-color: #FFFFFF !important; 
        color: #000000 !important; 
    }
    
    /* 5. NỀN BẢNG DỮ LIỆU CHÍNH (PIPELINE) -> TRẮNG MỊN */
    div[data-testid="stDataFrame"] > div:last-child,
    div[data-testid="stDataFrame"] { 
        background-color: #FFFFFF !important; 
    }
    .stDataFrame .data-cell {
        background-color: #FFFFFF !important; 
        color: #000000 !important;
    }
    
    /* **FIX: MÀU CHỮ TRONG CÁC Ô DATA EDITOR** */
    .stDataFrame .data-cell > div, 
    .stDataFrame .data-cell span,
    .stDataFrame .data-cell p,
    .stDataFrame .data-cell input,
    .stDataFrame .st-bh,
    .stDataFrame .st-br
    { 
        color: #000000 !important; 
        -webkit-text-fill-color: #000000 !important; 
    }
    .stDataFrame .data-cell input {
        background-color: #EBF5FB !important;
        color: #000000 !important;
        -webkit-text-fill-color: #000000 !important; 
    }
    
    /* --- V7.33.4: FINAL COLOR INJECTION VÀ WIDTH FIX CHO CỘT QUAN TRỌNG (NHẮM MỤC TIÊU SÂU HƠN) --- */
    /* Chú ý: Thứ tự cột VIEW mode: NAME (1), Cellphone (2), Số Tiệm (3), NOTE (4), 
       STATUS_SHORT (5) [GĐ (%)], TAM_LY_SHORT (6) [Tâm Lý].
    */
    
    /* 1. NOTE COLUMN (Cột thứ 4) - Xanh Mây Nhạt */
    div[data-testid="stDataFrame"] > div:nth-child(1) > div > div > div:nth-child(4) .data-cell {
        background-color: #E5F7FF !important; /* Xanh Mây */
        min-width: 250px !important; /* Tăng độ rộng cho NOTE */
        max-width: 300px !important;
    }
    
    /* 2. STATUS SHORT (GĐ %) COLUMN (Cột thứ 5) - Hồng Đào Nhạt */
    div[data-testid="stDataFrame"] > div:nth-child(1) > div > div > div:nth-child(5) .data-cell {
        background-color: #FCEEEA !important; /* Hồng Đào Nhạt */
        min-width: 70px !important; 
        max-width: 70px !important;
    }
    div[data-testid="stDataFrame"] > div:nth-child(1) > div > div > div:nth-child(5) {
         min-width: 70px !important; 
         max-width: 70px !important;
    }
    
    /* 3. TAM LY SHORT (Tâm Lý) COLUMN (Cột thứ 6) - Vàng Kem Nhạt */
    div[data-testid="stDataFrame"] > div:nth-child(1) > div > div > div:nth-child(6) .data-cell {
        background-color: #FFFBE5 !important; /* Vàng Kem Nhạt */
        min-width: 100px !important; 
        max-width: 100px !important;
    }
    div[data-testid="stDataFrame"] > div:nth-child(1) > div > div > div:nth-child(6) {
         min-width: 100px !important; 
         max-width: 100px !important;
    }
    
    /* 4. Cellphone Column (Cột thứ 2) - Rộng 100px */
    div[data-testid="stDataFrame"] > div:nth-child(1) > div > div > div:nth-child(2) .data-cell {
        min-width: 100px !important; 
        max-width: 100px !important;
    }
    div[data-testid="stDataFrame"] > div:nth-child(1) > div > div > div:nth-child(2) {
        min-width: 100px !important; 
        max-width: 100px !important;
    }
    
    /* 5. Số Tiệm Column (Cột thứ 3) - Rộng 100px */
    div[data-testid="stDataFrame"] > div:nth-child(1) > div > div > div:nth-child(3) .data-cell { 
        min-width: 100px !important; 
        max-width: 100px !important;
    }
    div[data-testid="stDataFrame"] > div:nth-child(1) > div > div > div:nth-child(3) {
        min-width: 100px !important; 
        max-width: 100px !important;
    }
    
    /* --- END COLOR & WIDTH CODING --- */


    
    /* 6. HEADER BẢNG -> XANH BĂNG */
    .stDataFrame > div > div:first-child > div {
        background-color: #EBF5FB !important; 
        border-bottom: 3px solid #D35400 !important; 
    }
    .stDataFrame .col-header-row .data-cell { 
        background-color: #EBF5FB !important; 
        color: #000000 !important; 
        font-weight: bold !important;
    }
    .stDataFrame > div > div:first-child > div > div {
        color: #000000 !important; 
        font-weight: bold !important; 
    }
    
    /* 7. Tiêu đề chính */
    h1 { color: #D35400 !important; border-bottom: 2px solid #D35400; }
    
    /* 8. Khu vực Upload File (Lỗi màu xanh đậm/đen) -> XANH BĂNG */
    div[data-testid="stFileUploaderDropzone"] {
        background-color: #EBF5FB !important;
        border: 2px dashed #B0C4DE !important;
        color: #000000 !important;
    }
    div[data-testid="stFileUploaderDropzone"] p {
        color: #000000 !important;
    }
    
    /* 9. Nút Tải Xuống trong Sidebar (Export) */
    div[data-testid="stDownloadButton"] button {
        background-color: #FFFFFF !important; 
        color: #000000 !important; 
        border: 1px solid #D35400 !important; 
    }
    div[data-testid="stDownloadButton"] button * {
        color: #000000 !important;
    }
    
    /* 10. Nút CHẠY LẠI AI TỔNG HỢP & LƯU THAY ĐỔI (PRIMARY/SECONDARY) */
    div[data-testid="stButton"] button {
        color: white !important;
        border: none !important;
    }
    /* FIX: Nút Secondary (COPY kịch bản, LOG CALL TIME) trong Sidebar */
    /* Dùng selector cha là sidebar để đảm bảo ưu tiên */
    section[data-testid="stSidebar"] div[data-testid="stButton"] button[kind="secondary"] {
        background-color: #FFFFFF !important;
        color: #000000 !important; /* CHỮ ĐEN */
        border: 1px solid #D35400 !important;
    }
    /* FIX: Đảm bảo chữ bên trong nút secondary trong sidebar là màu đen */
    section[data-testid="stSidebar"] div[data-testid="stButton"] button[kind="secondary"] * {
        color: #000000 !important; /* CHỮ ĐEN */
    }
    
    /* Nút Secondary khu vực chính (Nếu có - VD: CHẠY LẠI AI TỔNG HỢP) */
    div[data-testid="stButton"] button[kind="secondary"] {
        background-color: #FFFFFF !important;
        color: #D35400 !important; 
        border: 1px solid #D35400 !important;
    }
    
</style>
""", unsafe_allow_html=True)


# =============================================================================
# 2. LOGIC XỬ LÝ (AI & DATA)
# =============================================================================

# --- HÀM MỚI: GỌI GPT ĐỂ PHÂN TÍCH DỮ LIỆU (Dùng litellm) ---
# CẬP NHẬT: TĂNG STOP_AFTER_ATTEMPT LÊN 5 VÀ TĂNG WAIT_RANDOM_EXPONENTIAL
@retry(wait=wait_random_exponential(min=4, max=30), stop=stop_after_attempt(5), 
       retry=retry_if_exception_type(Exception))
def call_gpt_analysis(note_content, current_status):
    if AI_CLIENT_STATUS is not True:
        # Nếu AI chưa sẵn sàng do lỗi API Key, ném ra lỗi đặc biệt để dùng Fallback
        raise Exception("AI_CLIENT_NOT_READY") 

    note_content = str(note_content).strip()
    if not note_content:
        return "KHÔNG GHI CHÚ", "KHÔNG GỢI Ý", "KHÔNG KỊCH BẢN" 

    json_schema_prompt = """
    {
        "PHAN_TICH_TAM_LY": "Phân tích tâm lý khách hàng (Sử dụng một trong các từ khóa: TÍN HIỆU MUA, NGHI NGỜ, TỪ CHỐI, TÀI CHÍNH, HỨNG THẤP, KHÔNG RÕ)",
        "GOI_Y_HANH_DONG": "Chiến lược hành động ngắn gọn cho Sale (VD: CHỐT ĐƠN NGAY, TẠO NIỀM TIN, HẠ PHÍ, FOLLOW SÂU)",
        "NOI_DUNG_TU_VAN": "Kịch bản/nội dung tư vấn mẫu (1-2 câu) ngắn gọn, chuyên nghiệp, ready-to-copy. Nếu không thể tạo kịch bản, trả về chuỗi: 'KHÔNG KỊCH BẢN ĐƯỢC TẠO.'"
    }
    """

    system_prompt = f"""
    Bạn là một trợ lý AI chuyên nghiệp tên GUS, chuyên phân tích CRM cho ngành bảo hiểm nhân thọ/IUL (Mỹ/Canada).
    Nhiệm vụ của bạn là đọc Ghi chú (NOTE) của Sale và đưa ra 3 kết quả phân tích.
    Phản hồi **BẮT BUỘC** phải là một đối tượng JSON HỢP LỆ theo cấu trúc sau (KHÔNG thêm bất kỳ văn bản giải thích nào bên ngoài JSON):
    {json_schema_prompt}
    Hiện tại Status cũ là: {current_status}. Dữ liệu đầu vào chỉ là Ghi chú của Sale.
    """
    
    try:
        response = completion(
            model=AI_MODEL,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Ghi chú khách hàng: {note_content}"}
            ],
            response_format={"type": "json_object"} # Litellm dùng response_format
        )
        
        json_response = response.choices[0].message.content
        result = json.loads(json_response)
        
        # Bắt lỗi nếu Kịch bản là chuỗi rỗng và thay bằng chuỗi đặc biệt
        script = result.get("NOI_DUNG_TU_VAN", "Lỗi AI").strip()
        if not script or script.upper() in ["CHƯA CÓ KỊCH BẢN/LỖI AI TẠO.", "LỖI AI"]:
             script = "CHƯA CÓ KỊCH BẢN/LỖI AI TẠO."

        return (
            result.get("PHAN_TICH_TAM_LY", "Lỗi AI"), 
            result.get("GOI_Y_HANH_DONG", "Lỗi AI"), 
            script
        )

    except Exception as e:
        error_msg = str(e)
        # Ném ra lỗi cụ thể để Streamlit hiển thị rõ
        if "Incorrect API key" in error_msg or "401" in error_msg:
             st.error("❌ Lỗi Xác thực (401): API Key không hợp lệ hoặc đã hết hạn. Vui lòng kiểm tra lại Key.")
             raise e 
        elif "rate limit" in error_msg or "429" in error_msg:
             st.error("❌ Lỗi Quá tải (429): Đã vượt quá giới hạn gọi API. Đang thử lại...")
             raise e 
        else:
             st.error(f"❌ Lỗi gọi API (Chi tiết): {error_msg}")
             raise e 


# --- HÀM TỰ ĐỘNG CẬP NHẬT STATUS CŨ (Fallback) ---
def run_gus_analysis_fallback(note, current_status):
    default_tam_ly = "KHÔNG RÕ"
    default_hanh_dong = "Gửi thông tin chung."
    default_script = "Chào Anh/Chị..."
    
    current_status_updated = current_status
    note_lower = note.lower()

    for status_name, keywords in STATUS_RULES:
        is_matched = False
        for kw in keywords:
            if kw.lower() in note_lower or (status_name == "Unidentified (10%)" and note.strip() == ''):
                current_status_updated = status_name
                is_matched = True
                break
        if is_matched: break
        
    return [
        current_status_updated,
        f"🔘 {default_tam_ly} (AUTO)",
        f"💬 {default_hanh_dong} (AUTO)",
        f"📝 {default_script} (AUTO)"
    ]

# --- HÀM CHÍNH GỌI TÍNH NĂNG AI/AUTO CHO TOÀN BỘ DF ---
def run_gus_ai_analysis(df, force_ai_run=False):
    if df.empty: return df

    ai_cols = ["PHÂN TÍCH TÂM LÝ (GUS)", "GỢI Ý HÀNH ĐỘNG (GUS)", "NỘI DUNG TƯ VẤN (COPY)"]
    for col in ai_cols:
        if col not in df.columns: df[col] = "🔘 CHƯA PHÂN TÍCH"
    if "Status" not in df.columns: df["Status"] = "Unidentified (10%)"
    if "NOTE" not in df.columns: df["NOTE"] = ""

    results = []
    is_ai_ready = (AI_CLIENT_STATUS is True)
    
    # Sử dụng st.empty để chứa thông báo lỗi/spinner khi chạy tổng hợp
    status_placeholder = st.empty()

    with status_placeholder.container():
         st.info("⏳ Đang chạy AI phân tích... Vui lòng không đóng cửa sổ này.")

    
    for index, row in df.iterrows():
        note = str(row.get('NOTE', '')).strip()
        current_status = str(row.get('Status', 'Unidentified (10%)')).strip()
        
        should_run_ai = (
            note != '' and 
            not current_status.startswith("Done") and 
            not current_status.startswith("Stop") and
            (
                force_ai_run or 
                "CHƯA PHÂN TÍCH" in str(row['PHÂN TÍCH TÂM LÝ (GUS)']).upper() or
                "AUTO" in str(row['PHÂN TÍCH TÂM LÝ (GUS)']).upper() 
            )
        )
        
        if is_ai_ready and should_run_ai:
            try:
                # Cập nhật thông báo đang chạy theo từng khách hàng
                with status_placeholder.container():
                     st.info(f"⏳ Đang phân tích: {row.get('NAME', 'Khách hàng ẩn danh')}...")
                     
                tam_ly, hanh_dong, script = call_gpt_analysis(note, current_status)
                
                # Logic cập nhật Status tự động dựa trên kết quả AI
                new_status = current_status
                if "TÍN HIỆU MUA" in tam_ly.upper():
                    new_status = "Hot Interest (85%)"
                elif "TỪ CHỐI" in tam_ly.upper():
                    new_status = "Stop (0%)"
                elif "TÀI CHÍNH" in tam_ly.upper() or "NGHI NGỜ" in tam_ly.upper():
                    new_status = "Interest (75%)"
                elif "KHÔNG RÕ" in tam_ly.upper() or "HỨNG THẤP" in tam_ly.upper():
                    new_status = "Follow Up (50%)"

                results.append([
                    new_status,
                    f"🧠 {tam_ly} (AI)",
                    f"🎯 {hanh_dong} (AI)",
                    script
                ])
                # Xóa lỗi đã hiển thị nếu có
                status_placeholder.empty()

            except Exception as e:
                # Nếu AI thất bại sau tất cả các lần thử lại (do 401, hoặc lỗi API khác)
                if str(e) == "AI_CLIENT_NOT_READY":
                     st.toast("❌ Lỗi AI CRITICAL: API Key không hợp lệ. Dùng Fallback.", icon="🤖")
                else:
                     st.toast("❌ AI phân tích thất bại sau nhiều lần thử lại. Dùng Fallback.", icon="🤖")
                
                # Dùng Fallback
                status_upd, tam_ly_upd, hanh_dong_upd, script_upd = run_gus_analysis_fallback(note, current_status)
                results.append([status_upd, tam_ly_upd, hanh_dong_upd, script_upd])
            
        else:
            # Chạy logic Fallback (Tự động cập nhật Status bằng từ khóa) nếu AI không cần chạy
            status_upd, tam_ly_upd, hanh_dong_upd, script_upd = run_gus_analysis_fallback(note, current_status)
            
            final_status = row['Status']
            final_tam_ly = row['PHÂN TÍCH TÂM LÝ (GUS)'] 
            final_hanh_dong = row['GỢI Ý HÀNH ĐỘNG (GUS)']
            final_script = row['NỘI DUNG TƯ VẤN (COPY)']
            
            if "CHƯA PHÂN TÍCH" in str(row['PHÂN TÍCH TÂM LÝ (GUS)']).upper() or "AUTO" in str(row['PHÂN TÍCH TÂM LÝ (GUS)']).upper() or force_ai_run:
                 final_status = status_upd
                 final_tam_ly = tam_ly_upd
                 final_hanh_dong = hanh_dong_upd
                 final_script = script_upd
                
            results.append([
                final_status,
                final_tam_ly,
                final_hanh_dong,
                final_script
            ])

    df[['Status', "PHÂN TÍCH TÂM LÝ (GUS)", "GỢI Ý HÀNH ĐỘNG (GUS)", "NỘI DUNG TƯ VẤN (COPY)"]] = pd.DataFrame(results, index=df.index)
    
    status_placeholder.empty() # Xóa thông báo cuối cùng
    return df

# --- Hàm Load/Clean Data (FIXED: Thêm logic Unmerge và BỎ CACHE) ---

def clean_phone(phone_str):
    if pd.isna(phone_str) or phone_str == 'nan' or phone_str == '':
        return None
    return re.sub(r'[^0-9]+', '', str(phone_str))

def load_users():
    # CHÚ Ý: TRÊN CLOUD, CHÚNG TA CẦN users.xlsx ĐƯỢC COMMIT LÊN GITHUB
    try: return pd.read_excel("users.xlsx", engine="openpyxl")
    except: return pd.DataFrame()

def unmerge_excel_file(file_path):
    """Hàm này đọc file excel, hủy gộp tất cả các ô, và ghi đè lên file gốc."""
    # CHÚ Ý: KHI LÊN CLOUD, CHÚNG TA SẼ KHÔNG SỬ DỤNG data.xlsx NÀY NỮA, MÀ DÙNG FILE UPLOAD
    if 'STREAMLIT_SERVER_ADDRESS' in os.environ:
         # Nếu đang chạy trên Cloud, KHÔNG unmerge file gốc (file gốc không tồn tại hoặc không nên bị sửa)
         # Logic này chỉ cần thiết cho file upload
         return
    
    try:
        if not os.path.exists(file_path):
            return 
            
        wb = load_workbook(file_path)
        for sheet in wb.worksheets:
            if sheet.merged_cells:
                st.toast(f"⚠️ Đã phát hiện ô gộp trong sheet '{sheet.title}'. Đang tiến hành hủy gộp...", icon="🛠️")
                merged_cells_ranges = list(sheet.merged_cells.ranges)
                for merged_cell_range in merged_cells_ranges:
                    min_row, min_col, max_row, max_col = merged_cell_range.bounds
                    top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
                    sheet.unmerge_cells(str(merged_cell_range))
                    
                    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                         for cell in row:
                             if cell.value is None or str(cell.value).strip() == '':
                                  cell.value = top_left_cell_value
            
        wb.save(file_path)
        st.toast("✅ Đã hủy gộp ô (Unmerge Cells) thành công và lưu file.", icon="💾")
        
    except PermissionError:
        st.error(f"❌ Lỗi: File '{file_path}' đang được mở bởi Excel. Vui lòng đóng Excel và thử lại.")
    except Exception as e:
        st.warning(f"⚠️ Lỗi khi hủy gộp ô trong Excel: {e}. Vẫn tiếp tục đọc file.")


# --- ĐÃ XÓA @st.cache_data để FIX lỗi CacheReplayClosureError ---
def load_data():
    cols_standard = ['NAME', 'Cellphone', 'Số Tiệm', 'NOTE', 'Status', 'ASSIGNED', 'LAST_CONTACT_DATE', 'LAST_CALL_DATETIME', "PHÂN TÍCH TÂM LÝ (GUS)", "GỢI Ý HÀNH ĐỘNG (GUS)", "NỘI DUNG TƯ VẤN (COPY)"]
    
    try:
        # Nếu đang chạy trên Cloud, file data.xlsx không tồn tại hoặc không nên được đọc/ghi trực tiếp
        # Nếu chạy local, nó sẽ cố gắng đọc file data.xlsx
        
        # 1. FIX: UNMERGE CELLS trước khi đọc DF (Chỉ áp dụng Local)
        unmerge_excel_file("data.xlsx") 

        # 2. Đọc DataFrame 
        df = pd.read_excel("data.xlsx", engine="openpyxl")
        df.columns = df.columns.str.strip()
        
        if 'STATUS' in df.columns and 'Status' not in df.columns:
            df.rename(columns={'STATUS': 'Status'}, inplace=True)
            
        for col in ['NAME', 'Cellphone', 'Số Tiệm', 'NOTE', 'Status', 'ASSIGNED', "PHÂN TÍCH TÂM LÝ (GUS)", "GỢI Ý HÀNH ĐỘNG (GUS)", "NỘI DUNG TƯ VẤN (COPY)"]:
            if col not in df.columns:
                df[col] = ''
            df[col] = df[col].astype(str).replace('nan', '', regex=True).str.strip()
        
        for col in ["PHÂN TÍCH TÂM LÝ (GUS)", "GỢI Ý HÀNH ĐỘNG (GUS)", "NỘI DUNG TƯ VẤN (COPY)"]:
             if col not in df.columns or df[col].astype(str).str.strip().eq('').all():
                  df[col] = "🔘 CHƯA PHÂN TÍCH"
        
        if 'Status' in df.columns and (df['Status'].empty or df['Status'].iloc[0] == ''):
            df['Status'] = "Unidentified (10%)"
            
        if 'LAST_CONTACT_DATE' not in df.columns:
            df['LAST_CONTACT_DATE'] = date.today()
        else:
            df['LAST_CONTACT_DATE'] = pd.to_datetime(df['LAST_CONTACT_DATE'], errors='coerce').dt.normalize().dt.date
            df['LAST_CONTACT_DATE'] = df['LAST_CONTACT_DATE'].fillna(date.today())
            
        if 'LAST_CALL_DATETIME' not in df.columns:
            df['LAST_CALL_DATETIME'] = pd.NaT 
        else:
            df['LAST_CALL_DATETIME'] = pd.to_datetime(df['LAST_CALL_DATETIME'], errors='coerce')

        df['CLEAN_PHONE'] = df['Cellphone'].apply(clean_phone)
        
        # Loại bỏ các cột link video cũ nếu chúng tồn tại trong DF
        df = df.drop(columns=[col for col in VIDEO_MENU_KEYS if col in df.columns], errors='ignore')
        
        # Thêm cột CLEAN_SHOP_PHONE để tránh lỗi nếu không có cột Số Tiệm
        if 'Số Tiệm' in df.columns:
            df['CLEAN_SHOP_PHONE'] = df['Số Tiệm'].apply(clean_phone)
        else:
            df['CLEAN_SHOP_PHONE'] = None

        return df
        
    except FileNotFoundError:
        # Nếu đang chạy trên Cloud, FileNotFoundError là bình thường
        st.warning(f"⚠️ File data.xlsx không tồn tại. Vui lòng nạp file mới trong mục Import.")
        cols_standard_clean = [col for col in cols_standard] # Chỉ giữ các cột standard
        return pd.DataFrame(columns=cols_standard_clean)
        
    except Exception as e: 
        st.error(f"❌ Lỗi đọc file data.xlsx: {e}. Vui lòng kiểm tra file excel hoặc nạp file mới.")
        cols_standard_clean = [col for col in cols_standard]
        return pd.DataFrame(columns=cols_standard_clean) 

def normalize_columns(df_input):
    df = df_input.copy()
    rename_map = {}
    for col in df.columns:
        col_lower = str(col).lower().strip()
        for standard, aliases in MAPPING_DICT.items():
            if col_lower == standard.lower() or col_lower in aliases:
                rename_map[col] = standard
                break
    if rename_map: df = df.rename(columns=rename_map)
    return df

# --- LEADERBOARD LOGIC (FIXED: Bắt lỗi kiểu dữ liệu) ---
def calculate_leaderboard(df_input):
    if df_input.empty or 'ASSIGNED' not in df_input.columns:
        return pd.DataFrame(columns=['ASSIGNED', 'Total Leads', 'Done Count', 'Hot Count', 'Closing Rate (%)']).set_index('ASSIGNED')

    df = df_input.copy()
    df['ASSIGNED'] = df['ASSIGNED'].astype(str).str.strip()
    df = df[df['ASSIGNED'] != '']
    
    if 'Status' in df.columns:
        df['Status'] = df['Status'].fillna('Unidentified (10%)').astype(str)
    else:
        return pd.DataFrame() 

    leaderboard_df = df.groupby('ASSIGNED').agg(
        Total_Leads=('NAME', 'size'),
        Done_Count=('Status', lambda x: (x == "Done (100%)").sum()), 
        Hot_Count=('Status', lambda x: (x == "Hot Interest (85%)").sum())
    ).reset_index()
    
    # FIX: Chuyển đổi sang integer để tránh lỗi TypeError: Expected numeric dtype, got object instead.
    leaderboard_df['Total_Leads'] = pd.to_numeric(leaderboard_df['Total_Leads'], errors='coerce').fillna(0).astype(int)
    leaderboard_df['Done_Count'] = pd.to_numeric(leaderboard_df['Done_Count'], errors='coerce').fillna(0).astype(int)
    leaderboard_df['Hot_Count'] = pd.to_numeric(leaderboard_df['Hot_Count'], errors='coerce').fillna(0).astype(int)

    leaderboard_df['Closing Rate (%)'] = np.where(
        leaderboard_df['Total_Leads'] > 0,
        ((leaderboard_df['Done_Count'].astype(float) / leaderboard_df['Total_Leads'].astype(float)) * 100).round(1),
        0.0
    )

    leaderboard_df = leaderboard_df.sort_values(
        by=['Done_Count', 'Hot_Count', 'Total_Leads'], 
        ascending=[False, False, False]
    ).reset_index(drop=True)

    leaderboard_df.rename(columns={
        'Total_Leads': 'Total Leads',
        'Done_Count': 'Done Count',
        'Hot_Count': 'Hot Count'
    }, inplace=True)

    leaderboard_df.index.name = None
    leaderboard_df.insert(0, 'RANK 🏅', range(1, len(leaderboard_df) + 1))
    
    return leaderboard_df

# --- HÀM EXPORT DATA (Đã fix lỗi xlsxwriter) ---
def to_excel(df):
    """Convert DataFrame to Excel format in memory."""
    import io # Đảm bảo io được import trong đây
    output = io.BytesIO()
    # SỬ DỤNG 'openpyxl' LÀM ENGINE DỰ PHÒNG 
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    # Loại bỏ các cột không cần thiết cho Export
    cols_to_drop = [
        "CLEAN_PHONE", "LAST_CALL_DATETIME", "CLEAN_SHOP_PHONE",
        "PHÂN TÍCH TÂM LÝ (GUS)", "GỢI Ý HÀNH ĐỘNG (GUS)", "NỘI DUNG TƯ VẤN (COPY)",
        "STATUS_SHORT", "TAM_LY_SHORT", "VIDEO_GUIDE" 
    ]
    # Loại bỏ các cột link cũ nếu chúng vô tình được giữ lại
    all_cols_to_drop = cols_to_drop + VIDEO_MENU_KEYS
    df_clean = df.drop(columns=[c for c in all_cols_to_drop if c in df.columns], errors='ignore')

    df_clean.to_excel(writer, index=False, sheet_name='3M_Gus_Export')
    writer.close() 
    processed_data = output.getvalue()
    return processed_data


# =============================================================================
# 3. GIAO DIỆN NGƯỜI DÙNG
# =============================================================================

# --- INITIALIZE SESSION STATE ---
if "logged_in" not in st.session_state: st.session_state.logged_in = False
if "user_info" not in st.session_state: st.session_state.user_info = {}
if 'edit_mode' not in st.session_state: st.session_state.edit_mode = False
if 'edited_df' not in st.session_state: st.session_state.edited_df = pd.DataFrame() 


# FIX: Load data không dùng cache
st.session_state['original_df'] = load_data()

if st.session_state.edited_df.empty or st.session_state.edited_df.shape != st.session_state.original_df.shape:
    st.session_state.edited_df = st.session_state.original_df.copy()


def login_ui():
    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown("<h1 style='text-align: center; border: none; color: #D35400;'>3M-Gus CRM</h1>", unsafe_allow_html=True) # ĐÃ ĐỔI TÊN Ở ĐÂY
        
        # --- HIỂN THỊ LỖI AI NẾU CÓ ---
        if AI_ERROR:
             st.error(AI_ERROR)
        # -----------------------------
        
        with st.form("login_form"):
            u = st.text_input("Tên đăng nhập")
            p = st.text_input("Mật khẩu", type="password")
            btn = st.form_submit_button("ĐĂNG NHẬP", type="primary", use_container_width=True)
            
            if btn:
                users = load_users()
                if not users.empty:
                    user = users[(users['username'].astype(str)==str(u).strip()) & (users['password'].astype(str)==str(p).strip())]
                    if not user.empty:
                        st.session_state.logged_in = True
                        st.session_state.user_info = user.iloc[0].to_dict()
                        st.session_state['original_df'] = load_data()
                        st.session_state.edited_df = st.session_state.original_df.copy() 
                        st.rerun()
                    else: st.error("❌ Sai thông tin!")
                else: st.error("⚠️ Chưa có file users.xlsx! Vui lòng nạp file.")

def logout():
    st.session_state.logged_in = False
    st.rerun()

# --- HÀM HIỂN THỊ AI INSIGHT PANEL ---
def display_ai_insight_panel(row, original_index):
    st.markdown("<h4 style='color: #FFFFFF; border-bottom: 1px solid #FFFFFF;'>🧠 AI INSIGHT GUS</h4>", unsafe_allow_html=True) # Đặt màu trắng cho tiêu đề sidebar
    
    tam_ly = str(row.get("PHÂN TÍCH TÂM LÝ (GUS)", "Chưa phân tích")).strip()
    hanh_dong = str(row.get("GỢI Ý HÀNH ĐỘNG (GUS)", "Không có gợi ý")).strip()
    script = str(row.get("NỘI DUNG TƯ VẤN (COPY)", "CHƯA CÓ KỊCH BẢN/LỖI AI TẠO.")).strip() # Sửa mặc định
    note_content = str(row.get("NOTE", "")).strip()

    # 1. PHÂN TÍCH TÂM LÝ
    st.markdown(f"**1. Trạng thái Tâm lý:**")
    color_map = {
        "TỪ CHỐI": "error", "NGHI NGỜ": "warning", 
        "TÍN HIỆU MUA": "success", "TÀI CHÍNH": "info",
        "HỨNG THẤP": "info", "KHÔNG RÕ": "info",
        "KHÔNG GHI CHÚ": "info"
    }
    tam_ly_key = next((key for key in color_map if key.upper() in tam_ly.upper()), "info") 
    
    if "AI" in tam_ly.upper() or "AUTO" in tam_ly.upper():
        display_tam_ly = re.sub(r'\s+\(AI\)|\s+\(AUTO\)|\s+\(GUS\)|🔘|🧠|💬|🎯|📝', '', tam_ly).strip()
        if tam_ly_key == "success":
            st.success(f"**🧠 {display_tam_ly}**") 
        elif tam_ly_key == "error":
            st.error(f"**⛔ {display_tam_ly}**") 
        else:
            st.info(f"**💡 {display_tam_ly}**")
    else: 
        st.warning(f"**⚠️ {tam_ly}**")
        
    # 2. GỢI Ý HÀNH ĐỘNG
    st.markdown(f"**2. Gợi ý Hành động:**")
    display_hanh_dong = re.sub(r'\s+\(AI\)|\s+\(AUTO\)|\s+\(GUS\)|🔘|🧠|💬|🎯|📝', '', hanh_dong).strip()
    st.success(f"**{display_hanh_dong}**", icon="🎯") 

    # 3. KỊCH BẢN TƯ VẤN (Dùng Expander và Nút Copy)
    st.markdown("---")
    with st.expander("📄 Kịch bản Tư vấn Nhanh (Click để mở)"):
        # Cập nhật logic kiểm tra kịch bản
        if script and script.upper() not in ["CHƯA CÓ KỊCH BẢN/LỖI AI TẠO.", "LỖI AI", "KHÔNG KỊCH BẢN"]:
            # TEXT AREA CÓ THỂ BỊ LỖI MÀU TRẮNG TRÊN NỀN TRẮNG CỦA EXPANDER KHI KHÔNG DÙNG DARK MODE
            st.text_area("Kịch bản:", script, height=150, key="ai_script_text", disabled=True)
            if st.button("📋 Xác nhận đã COPY kịch bản", use_container_width=True, type="secondary"):
                st.toast("✅ Đã COPY kịch bản vào bộ nhớ đệm (Ctrl+C). Giờ Sếp có thể dán!", icon="📋")
        else:
            st.warning("Chưa có kịch bản được tạo cho Note này (hoặc Ghi chú quá ngắn).")
            
    # --- NÚT KÍCH HOẠT PHÂN TÍCH AI (FORCE DISPLAY) ---
    st.markdown("---")
    
    if not AI_CLIENT_STATUS:
         st.warning(f"Tính năng AI Tắt: {AI_ERROR}")
    else:
        # Điều kiện hiển thị thông báo thay vì ẩn nút
        if not note_content:
             st.info("⚠️ Vui lòng nhập Ghi chú (NOTE) để AI phân tích chính xác.")
        if str(row.get('Status', '')).startswith(("Done", "Stop")):
             st.info("Khách hàng đã ở trạng thái Done/Stop. AI phân tích không cần thiết.")
        
        is_analyzed_by_ai = str(tam_ly).startswith("🧠")
        button_label = "🤖 CHẠY PHÂN TÍCH AI (1 LẦN)" if not is_analyzed_by_ai else "🔄 CHẠY LẠI AI PHÂN TÍCH"

        if st.button(button_label, use_container_width=True, type="primary"):
             with st.spinner("⏳ AI đang phân tích lại ghi chú..."):
                 try:
                     tam_ly, hanh_dong, script_result = call_gpt_analysis(note_content, str(row.get('Status')))
                     
                     if tam_ly != "AI_CLIENT_NOT_READY" and tam_ly != "KHÔNG GHI CHÚ":
                         new_status = str(row.get('Status'))
                         if "TÍN HIỆU MUA" in tam_ly.upper():
                             new_status = "Hot Interest (85%)"
                         elif "TỪ CHỐI" in tam_ly.upper():
                             new_status = "Stop (0%)"
                         elif "TÀI CHÍNH" in tam_ly.upper() or "NGHI NGỜ" in tam_ly.upper():
                             new_status = "Interest (75%)"
                         elif "KHÔNG RÕ" in tam_ly.upper() or "HỨNG THẤP" in tam_ly.upper():
                             new_status = "Follow Up (50%)"
                         
                         
                         st.session_state.edited_df.loc[original_index, 'Status'] = new_status
                         st.session_state.edited_df.loc[original_index, 'PHÂN TÍCH TÂM LÝ (GUS)'] = f"🧠 {tam_ly} (AI)"
                         st.session_state.edited_df.loc[original_index, 'GỢI Ý HÀNH ĐỘNG (GUS)'] = f"🎯 {hanh_dong} (AI)"
                         st.session_state.edited_df.loc[original_index, 'NỘI DUNG TƯ VẤN (COPY)'] = script_result # Lưu script đã được làm sạch/kiểm tra
                            
                         save_dataframe_changes(st.session_state.edited_df)
                         # Sau khi lưu, buộc phải tải lại data mới (không dùng cache)
                         st.session_state.original_df = load_data()
                         st.session_state.edited_df = st.session_state.original_df.copy()
                         st.rerun()
                     else:
                         st.warning("Ghi chú trống hoặc AI không thể phân tích.")

                 except Exception as e:
                      # Bắt lỗi AI CRITICAL (như API Key, Rate Limit)
                      if str(e) != "AI_CLIENT_NOT_READY":
                        st.error(f"Lỗi CRITICAL khi gọi AI: {type(e).__name__}.")

# --- HÀM HIỂN THỊ POP-UP VIDEO (Mục tiêu 1.5) ---
def display_video_popup_panel():
    st.markdown("---")
    st.markdown("<h4 style='color: #FFFFFF;'>▶️ VIDEO TÀI LIỆU CHUYÊN SÂU</h4>", unsafe_allow_html=True)
    
    # Tạo các nút LinkButton trong Sidebar
    for name, url in MENU_VIDEO.items():
         # Đã chuyển sang PRIMARY để fix lỗi màu chữ (V7.31)
         st.link_button(
             label=f"🎬 {name}", 
             url=url, 
             type="primary", 
             use_container_width=True
         )
    st.caption("Các nút này mở Video trong tab mới.")
    

# --- MAIN APP LOGIC ---
def main_app():
    user = st.session_state.user_info
    df_current = st.session_state['original_df'].copy()
    
    with st.sidebar:
        st.markdown(f"## 👤 {user['name']}")
        st.caption(f"Role: {str(user['role']).upper()}")
        st.markdown("---")
        menu = st.radio("MENU ĐIỀU HƯỚNG", ["📊 Dashboard", "📇 Pipeline Khách Hàng", "📥 Import & AI Phân Tích", "⚙️ Cài Đặt Hệ Thống"])
        st.markdown("---")
        
        if menu == "📇 Pipeline Khách Hàng":
            st.markdown("### 📞 GỌI ĐIỆN VÀ CHỈNH SỬA")
            
            if not st.session_state['original_df'].empty and 'NAME' in st.session_state['original_df'].columns:
                df_valid_contacts = st.session_state['original_df'][
                    (st.session_state['original_df']['NAME'].astype(str).str.strip() != '') | 
                    (st.session_state['original_df']['Cellphone'].astype(str).str.strip() != '')
                ].copy()
                
                df_valid_contacts['NAME'] = df_valid_contacts['NAME'].fillna('').astype(str).str.strip()
                df_valid_contacts['Cellphone'] = df_valid_contacts['Cellphone'].fillna('').astype(str).str.strip()

                df_valid_contacts['DISPLAY'] = df_valid_contacts.apply(
                    lambda row: f"{row['NAME']} ({row['Cellphone']})" if row['NAME'] and row['Cellphone'] 
                                 else (row['NAME'] if row['NAME'] else row['Cellphone']), 
                    axis=1
                )
                
                customer_options = ['--- Chọn khách hàng/SĐT ---'] + sorted(df_valid_contacts['DISPLAY'].tolist())
                
                # BẮT BUỘC SELECT BOX CHO CALL DISPLAY VỀ MÀU SÁNG
                selected_display = st.selectbox(
                    "Chọn Khách Hàng (Tên/SĐT)", 
                    customer_options,
                    key='call_select'
                )
                
                selected_row = None
                
                if selected_display != '--- Chọn khách hàng/SĐT ---':
                    try:
                        selected_row = df_valid_contacts[df_valid_contacts['DISPLAY'] == selected_display].iloc[0]
                        row = selected_row.to_dict()
                        original_index = selected_row.name 

                        kh_phone_clean = clean_phone(row.get('Cellphone'))
                        tiem_phone_clean = clean_phone(row.get('Số Tiệm'))
                        
                        st.caption(f"Đang chọn: **{row.get('NAME', 'N/A')}**")
                        
                        col_kh, col_tiem = st.columns(2)
                        
                        if kh_phone_clean:
                            with col_kh:
                                st.markdown(
                                    f'<a href="tel:+1{kh_phone_clean}" target="_self"><button style="background-color: #58D68D; color: white; border-radius: 5px; border: none; padding: 5px 10px; cursor: pointer;">📞 GỌI KH</button></a>',
                                    unsafe_allow_html=True
                                )
                                st.caption(f"SĐT KH: {row.get('Cellphone', 'N/A')}")
                        
                        if tiem_phone_clean:
                            with col_tiem:
                                st.markdown(
                                    f'<a href="tel:+1{tiem_phone_clean}" target="_self"><button style="background-color: #5DADE2; color: white; border-radius: 5px; border: none; padding: 5px 10px; cursor: pointer;">📞 GỌI TIỆM</button></a>',
                                    unsafe_allow_html=True
                                )
                                st.caption(f"SĐT Tiệm: {row.get('Số Tiệm', 'N/A')}")
                        
                        display_ai_insight_panel(row, original_index)

                        if kh_phone_clean or tiem_phone_clean:
                            st.markdown("---")
                            # NÚT ĐÃ GỌI
                            if st.button("📞 XÁC NHẬN ĐÃ GỌI (Log Call Time)", key='confirm_call_log', use_container_width=True, type="secondary"):
                                
                                st.session_state.edited_df.loc[original_index, 'LAST_CONTACT_DATE'] = date.today()
                                st.session_state.edited_df.loc[original_index, 'LAST_CALL_DATETIME'] = datetime.now()
                                
                                save_dataframe_changes(st.session_state.edited_df)
                                # Sau khi lưu, buộc phải tải lại data mới (không dùng cache)
                                st.session_state.original_df = load_data()
                                st.session_state.edited_df = st.session_state.original_df.copy()
                                st.rerun()

                        # --- HIỂN THỊ POP-UP VIDEO PANEL ---
                        display_video_popup_panel()


                    except Exception as e:
                        st.error("Lỗi khi tải thông tin SĐT.")


            st.markdown("---")
            
            if st.session_state.edit_mode:
                if st.button("🔴 TẮT CHẾ ĐỘ CHỈNH SỬA", use_container_width=True, type="secondary"):
                    st.session_state.edit_mode = False
                    st.rerun()
            else:
                if st.button("🟢 MỞ CHẾ ĐỘ CHỈNH SỬA", use_container_width=True, type="primary"):
                    st.session_state.edit_mode = True
                    st.rerun()
            
            # --- TÍNH NĂNG EXPORT ---
            st.markdown("---")
            st.markdown("### 📥 EXPORT DỮ LIỆU")
            
            if 'df_display' in st.session_state and not st.session_state.df_display.empty:
                df_export = st.session_state.df_display.copy()
            else:
                df_export = st.session_state.edited_df.copy()
            
            # Nút Tải Xuống
            st.download_button(
                label="⬇️ Tải Xuống File Excel Đã Lọc",
                data=to_excel(df_export),
                file_name=f"3M_Gus_Export_{user['username']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            # -------------------------------

        st.markdown("---")
        if st.button("Đăng Xuất", use_container_width=True): logout()
        
    
    if menu == "📊 Dashboard":
        st.title("📊 DASHBOARD TỔNG QUAN")
        
        if df_current.empty: 
            st.warning("Dữ liệu trống. Vui lòng vào mục Import để nạp file.")
            return

        total_leads = len(df_current)
        today = date.today() 
        
        df_current['LAST_CONTACT_DATE_COMPARE'] = pd.to_datetime(df_current['LAST_CONTACT_DATE'], errors='coerce').dt.date
        df_current['Status_Percent'] = df_current['Status'].astype(str).str.extract(r'\((\d+)%\)').astype(float).fillna(0)
        
        leads_to_call = df_current[
            (df_current['LAST_CONTACT_DATE_COMPARE'] < today) & 
            (df_current['Status_Percent'] < 100)
        ]
        count_to_call = len(leads_to_call)
        
        ai_col = df_current['PHÂN TÍCH TÂM LÝ (GUS)'].fillna('🔘 CHƯA PHÂN TÍCH').astype(str)
        hot_leads = df_current[df_current['Status'].astype(str) == "Hot Interest (85%)"]
        count_hot = len(hot_leads)
        done_leads = df_current[df_current['Status'].astype(str) == "Done (100%)"]
        count_done = len(done_leads)
        stop_leads = df_current[df_current['Status'].astype(str) == "Stop (0%)"]
        count_stop = len(stop_leads)
        
        
        st.markdown("### 📈 Chỉ số Hiệu suất Chính (KPIs)")
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric(label="Tổng số Khách Hàng", value=total_leads, delta_color="off")
        with col2:
            st.metric(
                label="Khách Cần Gọi Lại 📞", 
                value=count_to_call, 
                delta=f"{round(count_to_call / total_leads * 100) if total_leads > 0 else 0}%",
                delta_color="inverse"
            )
        with col3:
            st.metric(
                label="Khách DONE ✅", 
                value=count_done, 
                delta=f"+{round(count_done / total_leads * 100) if total_leads > 0 else 0}%",
                delta_color="normal"
            )
        with col4:
             st.metric(
                label="Khách STOP/TỪ CHỐI ⛔", 
                value=count_stop, 
                delta=f"-{round(count_stop / total_leads * 100) if total_leads > 0 else 0}%",
                delta_color="inverse"
            )

        st.markdown("---")
        
        st.markdown("### 📊 Phân tích Dữ liệu")
        chart1, chart2 = st.columns(2)
        
        status_counts = df_current['Status'].value_counts().reset_index()
        status_counts.columns = ['Status', 'Count']
        
        fig_status = px.pie(
            status_counts, 
            values='Count', 
            names='Status', 
            title='Phân bổ Khách Hàng theo Giai đoạn Bán hàng (%)',
            hole=.3,
            color_discrete_sequence=px.colors.sequential.RdBu
        )
        fig_status.update_traces(textposition='inside', textinfo='percent+label')
        
        with chart1:
            st.plotly_chart(fig_status, use_container_width=True)
            
        ai_counts = ai_col.value_counts().reset_index()
        ai_counts.columns = ['AI_Analysis', 'Count']
        
        fig_ai = px.bar(
            ai_counts,
            x='AI_Analysis', 
            y='Count', 
            title='Phân tích Tâm lý Khách Hàng (AI)',
            color='AI_Analysis',
            color_discrete_sequence=px.colors.qualitative.Bold
        )
        fig_ai.update_layout(xaxis_title='Tâm lý Khách Hàng (GUS)', yaxis_title='Số Lượng')

        with chart2:
            st.plotly_chart(fig_ai, use_container_width=True)

        st.markdown("---")
        st.markdown("### 🏅 Bảng Xếp Hạng Thi Đua (Leaderboard)")
        
        leaderboard_data = calculate_leaderboard(df_current)
        
        if not leaderboard_data.empty:
            styled_leaderboard = leaderboard_data.style.format({
                'Closing Rate (%)': "{:.1f}%"
            })
            
            st.dataframe(
                styled_leaderboard,
                use_container_width=True,
                hide_index=True,
                column_order=['RANK 🏅', 'ASSIGNED', 'Done Count', 'Hot Count', 'Total Leads', 'Closing Rate (%)']
            )
        else:
            st.info("Chưa có khách hàng được gán cho Sale để tính toán Leaderboard.")

    elif menu == "📇 Pipeline Khách Hàng":
        st.title("📇 PIPELINE KHÁCH HÀNG")
        
        col_search, col_filter = st.columns([3, 1])

        with col_search: search = st.text_input("🔍 Tìm kiếm nhanh", placeholder="Nhập tên, SĐT...")
        with col_filter: 
            filter_mode = st.selectbox("Lọc dữ liệu", [
                "Tất cả", 
                "Khách Cần Gọi Lại 📞 (Dưới 24H)", 
                "Khách Hàng LẠNH 🧊 (Trên 14 Ngày)",
                "Khách DONE ✅",
                "Khách HOT 🔥 (85%)",
                "Khách Tương Tác Mới ✨" 
            ])

        if not df_current.empty:
            df_show = st.session_state.edited_df.copy() 
            current_datetime = datetime.now()
            today = date.today()
            
            if user['role'] != 'admin' and "ASSIGNED" in df_show.columns:
                df_show = df_show[df_show['ASSIGNED'].astype(str) == str(user['username'])]
            
            if 'LAST_CONTACT_DATE' in df_show.columns:
                df_show['LAST_CONTACT_DATE'] = pd.to_datetime(df_show['LAST_CONTACT_DATE'], errors='coerce').dt.normalize().dt.date
                df_show = df_show.dropna(subset=['LAST_CONTACT_DATE'])
            
            if "Status" in df_show.columns:
                 status_filter_col = df_show["Status"].fillna('Unidentified (10%)').astype(str)
            else:
                 status_filter_col = pd.Series([""] * len(df_show.index)) 
                 
            if filter_mode == "Khách DONE ✅":
                df_show = df_show[status_filter_col.str.contains("Done \(100%\)", regex=True)]
            
            elif filter_mode == "Khách HOT 🔥 (85%)":
                df_show = df_show[status_filter_col.str.contains("Hot Interest \(85%\)", regex=True)]

            elif filter_mode == "Khách Cần Gọi Lại 📞 (Dưới 24H)" and 'LAST_CALL_DATETIME' in df_show.columns:
                time_24h_ago = current_datetime - timedelta(hours=24)
                df_show['LAST_CALL_DATETIME'] = pd.to_datetime(df_show['LAST_CALL_DATETIME'], errors='coerce')
                df_show = df_show[
                    (df_show['LAST_CALL_DATETIME'].isna()) | 
                    (df_show['LAST_CALL_DATETIME'] < time_24h_ago) 
                ]
                df_show = df_show[~status_filter_col.str.contains("Done \(100%\)", regex=True)]

            elif filter_mode == "Khách Hàng LẠNH 🧊 (Trên 14 Ngày)" and 'LAST_CONTACT_DATE' in df_show.columns:
                date_14_days_ago = today - timedelta(days=14)
                df_show = df_show[
                    (df_show['LAST_CONTACT_DATE'] < date_14_days_ago) & 
                    (~status_filter_col.str.contains("Done \(100%\)", regex=True))
                ]
            
            elif filter_mode == "Khách Tương Tác Mới ✨" and 'LAST_CONTACT_DATE' in df_show.columns:
                 df_show = df_show[df_show['LAST_CONTACT_DATE'] == today]
            
            if search:
                mask = df_show.astype(str).apply(lambda x: x.str.contains(search, case=False, na=False)).any(axis=1)
                df_show = df_show[mask]

            # GÁN CỘT HIỂN THỊ NGẮN GỌN (Mục tiêu 1.1)
            # STATUS_SHORT: Chỉ lấy %
            df_show['STATUS_SHORT'] = df_show['Status'].astype(str).str.extract(r'\((\d+)%\)').fillna('0%') + ' %'
            # TAM_LY_SHORT: Bỏ các ký hiệu/nguồn
            df_show['TAM_LY_SHORT'] = df_show['PHÂN TÍCH TÂM LÝ (GUS)'].astype(str).str.replace(r'\(AI\)|\(AUTO\)|🔘|🧠|💬|🎯|📝', '', regex=True).str.strip()
            
            # Gán df_show vào session state để dùng cho Export
            st.session_state.df_display = df_show.copy()

            cols_to_drop = [
                "GỢI Ý HÀNH ĐỘNG (GUS)", 
                "NỘI DUNG TƯ VẤN (COPY)", 
                "ASSIGNED",
                "LAST_CONTACT_DATE",
                "CLEAN_PHONE",
                "LAST_CALL_DATETIME",
                "CLEAN_SHOP_PHONE",
                "PHÂN TÍCH TÂM LÝ (GUS)", # Ẩn cột cũ
                "Status", # Ẩn cột cũ
                "VIDEO_GUIDE" # Chắc chắn loại bỏ
            ]
            
            EDITABLE_COLS = ['NAME', 'Cellphone', 'Số Tiệm', 'NOTE', 'Status']
            
            # Loại bỏ các cột dư thừa cho bảng hiển thị
            df_display = df_show.drop(columns=[c for c in cols_to_drop if c in df_show.columns], errors='ignore')
            
            # Khôi phục cột Status/PHÂN TÍCH TÂM LÝ cho việc chỉnh sửa (Nếu cần)
            if st.session_state.edit_mode:
                 if 'Status' in df_show.columns: df_display['Status'] = df_show['Status']
                 if 'PHÂN TÍCH TÂM LÝ (GUS)' in df_show.columns: df_display['PHÂN TÍCH TÂM LÝ (GUS)'] = df_show['PHÂN TÍCH TÂM LÝ (GUS)']

            uneditable_cols_in_display = [
                col for col in df_display.columns if col not in EDITABLE_COLS
            ]
            
            column_config_base = {
                "NAME": st.column_config.TextColumn("NAME", max_chars=100, width="medium", help="Click để chỉnh sửa Tên Khách Hàng."),
                "Cellphone": st.column_config.TextColumn("Cellphone", max_chars=20, width="small", help="Click để chỉnh sửa Số Điện Thoại Khách."),
                "Số Tiệm": st.column_config.TextColumn("Số Tiệm", max_chars=20, width="small", help="Click để chỉnh sửa Số Điện Thoại Tiệm."),
                
                # Cột Status dài (Chỉ hiện khi Edit Mode)
                "Status": st.column_config.SelectboxColumn("Status (Chi tiết)", options=STATUS_OPTIONS, required=True, width="small", help="Giai đoạn bán hàng.", disabled=(not st.session_state.edit_mode)),
                
                # Cột PHÂN TÍCH TÂM LÝ (Chỉ hiện khi Edit Mode)
                "PHÂN TÍCH TÂM LÝ (GUS)": st.column_config.TextColumn("Tâm Lý (Chi tiết)", width="small", disabled=True),
                
                "NOTE": st.column_config.TextColumn("NOTE", max_chars=300, width="medium", help="Ghi chú."),
                
                # Cột hiển thị ngắn gọn (Chỉ hiện khi View Mode)
                "STATUS_SHORT": st.column_config.TextColumn("GĐ (%)", width="small", disabled=True, help="Giai đoạn bán hàng (Phần trăm)."),
                "TAM_LY_SHORT": st.column_config.TextColumn("Tâm Lý", width="small", disabled=True, help="Tâm lý khách hàng (AI/Auto)."),
            }
            
            final_column_config = {k: v for k, v in column_config_base.items() if k in df_display.columns}

            # Xác định thứ tự cột dựa trên chế độ chỉnh sửa
            column_order = ['NAME', 'Cellphone', 'Số Tiệm', 'NOTE']
            
            if st.session_state.edit_mode:
                 # Khi Edit, hiện cột Status/Tâm Lý dài để có thể chỉnh sửa Status
                 column_order += ['Status', 'PHÂN TÍCH TÂM LÝ (GUS)'] 
            else:
                 # Khi View, hiện cột Status/Tâm Lý ngắn gọn
                 column_order += ['STATUS_SHORT', 'TAM_LY_SHORT']

            if st.session_state.edit_mode:
                st.warning("⚠️ Đang ở **CHẾ ĐỘ CHỈNH SỬA**. Sau khi sửa xong, nhấn **LƯU THAY ĐỔI**.")
                
                edited_df = st.data_editor(
                    df_display,
                    column_config=final_column_config, 
                    use_container_width=True,
                    height=600,
                    hide_index=True,
                    disabled=[col for col in uneditable_cols_in_display if col not in ['Status', 'PHÂN TÍCH TÂM LÝ (GUS)']], # Chỉ cho phép chỉnh sửa Status/NOTE
                    column_order=column_order,
                    key='editor_data'
                )
                
                last_edited_index = None
                if 'editor_data' in st.session_state and 'edited_rows' in st.session_state['editor_data']:
                    edited_rows = st.session_state['editor_data']['edited_rows']
                    if edited_rows:
                        last_edited_index_in_view = list(edited_rows.keys())[-1]
                        
                        if last_edited_index_in_view < len(edited_df.index):
                            original_index = edited_df.index[last_edited_index_in_view]
                            last_edited_index = original_index
                
                if last_edited_index is not None:
                    if last_edited_index in st.session_state.edited_df.index:
                        full_note = st.session_state.edited_df.loc[last_edited_index, 'NOTE']
                        customer_name = st.session_state.edited_df.loc[last_edited_index, 'NAME']

                        st.markdown("---")
                        with st.expander(f"📝 **Ghi chú chi tiết của {customer_name}** (Click để ẩn/hiện)", expanded=True):
                            st.markdown(f"**Tên Khách Hàng:** {customer_name}")
                            st.markdown(f"**Nội dung Note ĐẦY ĐỦ:**")
                            st.info(full_note)

                if st.button("💾 LƯU THAY ĐỔI", type="primary"):
                    if 'editor_data' in st.session_state:
                        
                        changes = st.session_state['editor_data']['edited_rows']
                        current_date = date.today() 
                        
                        for index, row_changes in changes.items():
                            original_index = edited_df.index[index]
                            
                            if 'LAST_CONTACT_DATE' in st.session_state.edited_df.columns:
                                st.session_state.edited_df.loc[original_index, 'LAST_CONTACT_DATE'] = current_date
                                
                            for col, new_value in row_changes.items():
                                if col == 'Status' or col == 'NOTE':
                                    st.session_state.edited_df.loc[original_index, col] = new_value
                                
                        save_dataframe_changes(st.session_state.edited_df) 
                        
                        # Sau khi lưu, buộc phải tải lại data mới (không dùng cache)
                        st.session_state.original_df = load_data() 
                        st.session_state.edited_df = st.session_state.original_df.copy()
                        st.success("✅ Đã lưu thay đổi thành công! Tải lại dữ liệu.")
                        st.rerun()

            else:
                st.dataframe(
                    df_display,
                    column_config=final_column_config, 
                    use_container_width=True,
                    height=600,
                    hide_index=True,
                    column_order=column_order
                )
        else: st.warning("Danh sách trống. Vui lòng vào mục Import để nạp file.")

    elif menu == "📥 Import & AI Phân Tích":
        st.title("📥 IMPORT & AI PHÂN TÍCH")
        
        st.markdown("---")
        if AI_CLIENT_STATUS:
            if st.button("🤖 CHẠY LẠI AI PHÂN TÍCH TỔNG HỢP", type="secondary"):
                 df_updated = run_gus_ai_analysis(st.session_state.edited_df.copy(), force_ai_run=True)
                 
                 st.session_state.edited_df = df_updated.copy()
                 save_dataframe_changes(st.session_state.edited_df)
                 
                 # Sau khi lưu, buộc phải tải lại data mới (không dùng cache)
                 st.session_state.original_df = load_data()
                 st.session_state.edited_df = st.session_state.original_df.copy()
                 st.success("✅ AI đã phân tích lại toàn bộ dữ liệu. Kiểm tra mục Pipeline Khách Hàng.")
                 st.balloons()
                 st.rerun()
        else:
             st.warning(f"Tính năng AI Tắt: {AI_ERROR}")
        st.markdown("---")
        
        st.info("Kéo thả file Excel vào đây. Hệ thống sẽ tự động cập nhật Giai đoạn bán hàng (Status) sau khi Import và phân tích AI.")
        up = st.file_uploader("Tải file Excel", type=['xlsx'])
        if up:
            try:
                temp_import_path = "temp_import.xlsx"
                with open(temp_import_path, "wb") as f:
                    f.write(up.getbuffer())
                
                # FIX V6.0: UNMERGE CELLS cho file Import tạm thời trước khi đọc DF
                unmerge_excel_file(temp_import_path) 
                
                df_new = normalize_columns(pd.read_excel(temp_import_path, engine="openpyxl"))
                os.remove(temp_import_path) 
                
            except Exception as e:
                st.error(f"❌ Lỗi đọc file Excel: {e}")
                if os.path.exists(temp_import_path):
                     os.remove(temp_import_path)
                return

            if st.button("🚀 KÍCH HOẠT XỬ LÝ IMPORT", type="primary"):
                # Không hiển thị spinner bên ngoài, spinner/status được quản lý bên trong run_gus_ai_analysis
                st.info("⏳ Đang Import, phân tích AI và chống trùng lặp... Vui lòng chờ cho đến khi nhận được thông báo HOÀN TẤT.")
                
                # Chạy AI khi Import
                df_analyzed = run_gus_ai_analysis(df_new, force_ai_run=False)
                
                # Logic chống trùng lặp và hợp nhất
                df_base = st.session_state['original_df'].drop(columns=[c for c in st.session_state['original_df'].columns if c in VIDEO_MENU_KEYS], errors='ignore')
                
                if not df_base.empty: 
                    cols_to_drop_from_base = [
                        c for c in df_base.columns 
                        if c in df_analyzed.columns and c not in ['NAME', 'Cellphone', 'Số Tiệm', 'LAST_CONTACT_DATE']
                    ]
                    df_base_clean = df_base.drop(columns=cols_to_drop_from_base, errors='ignore')
                    df_final = pd.concat([df_base_clean, df_analyzed], ignore_index=True)
                else: 
                    df_final = df_analyzed

                if 'NAME' in df_final.columns and 'Cellphone' in df_final.columns:
                        df_final['NAME_TEMP'] = df_final['NAME'].astype(str).fillna('').apply(lambda x: x if x.strip()!='' else f'NO_NAME_{np.random.randint(100000)}')
                        df_final['CELLPHONE_TEMP'] = df_final['Cellphone'].astype(str).fillna('').apply(lambda x: x if x.strip()!='' else f'NO_PHONE_{np.random.randint(100000)}')
                        
                        df_final = df_final.drop_duplicates(subset=['NAME_TEMP', 'CELLPHONE_TEMP'], keep='last')
                        
                        df_final = df_final.drop(columns=['NAME_TEMP', 'CELLPHONE_TEMP'])
                        
                if 'LAST_CONTACT_DATE' not in df_final.columns:
                        df_final['LAST_CONTACT_DATE'] = date.today()
                
                if 'LAST_CALL_DATETIME' not in df_final.columns:
                        df_final['LAST_CALL_DATETIME'] = pd.NaT

                df_final = run_gus_ai_analysis(df_final, force_ai_run=False) # Chạy lại lần nữa để update các cột AI sau khi hợp nhất

                save_dataframe_changes(df_final)
                
                # Sau khi lưu, buộc phải tải lại data mới (không dùng cache)
                st.session_state['original_df'] = load_data()
                st.session_state.edited_df = st.session_state.original_df.copy()
                
                st.success("✅ HOÀN TẤT! Đã Import, chống trùng lặp và AI đã phân tích. Vui lòng vào mục Pipeline Khách Hàng để kiểm tra.")
                st.balloons()
                st.rerun()

    elif menu == "⚙️ Cài Đặt Hệ Thống":
        st.title("⚙️ CÀI ĐẶT HỆ THỐNG")
        st.write("Phiên bản: 3M-Gus CRM v7.33.4 (Cloud Ready)")
        st.markdown("---")
        
        st.subheader("🛠️ Trạng thái AI Client")
        if AI_CLIENT_STATUS:
            st.success(f"✅ AI Client Đã Sẵn Sàng (Model: {AI_MODEL}).")
            st.info("API Key được đọc thành công từ Streamlit Secrets hoặc Biến môi trường.")
        else:
            st.error(f"❌ AI Client Bị Lỗi: {AI_ERROR}")
            st.warning("Vui lòng **đặt OPENAI_API_KEY** vào Streamlit Secrets (Cloud) hoặc Biến môi trường (Local).")


if st.session_state.logged_in: main_app()
else: login_ui()